#!/usr/bin/env python3
import argparse
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


def read_text(path_arg: Optional[str]) -> str:
    if path_arg:
        return Path(path_arg).read_text(encoding="utf-8")
    return sys.stdin.read()


def parse_entries(raw: str) -> list[dict]:
    blocks = re.findall(r"\[voice\]\[qa\]\s*\{(.*?)\}", raw, flags=re.S)
    entries: list[dict] = []

    for block in blocks:
        entry: dict = {}
        for key, str_val, num_val in re.findall(r"(\w+)\s*:\s*(?:'([^']*)'|(\d+))", block):
            if str_val != "":
                entry[key] = str_val
            elif num_val != "":
                entry[key] = int(num_val)
        if entry.get("sessionId"):
            entries.append(entry)
    return entries


def parse_session_date(session_id: str) -> str:
    match = re.search(r"(\d{12,})$", session_id or "")
    if not match:
        return datetime.now().strftime("%Y-%m-%d")
    ts_ms = int(match.group(1))
    dt = datetime.fromtimestamp(ts_ms / 1000, tz=timezone.utc).astimezone()
    return dt.strftime("%Y-%m-%d")


def split_target_words(words: str) -> list[str]:
    return [w.strip() for w in (words or "").split(",") if w.strip()][:4]


def next_empty_row(ws, start=2, end=16):
    for row in range(start, end + 1):
        if not ws.cell(row, 2).value:
            return row
    return None


def clear_manual_rows(ws, start=2, end=16):
    # Keep formula columns intact (K, M, N, S) and only clear manual input cells.
    manual_cols = [2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 16, 17, 18, 20]
    for row in range(start, end + 1):
        for col in manual_cols:
            ws.cell(row, col).value = None


def write_entries(ws, entries: list[dict], dry_run: bool = False, append_from_row: int = 2) -> int:
    written = 0
    preview_lines: list[str] = []
    search_start = append_from_row
    for entry in entries:
        row = next_empty_row(ws, start=search_start, end=16)
        if row is None:
            break

        words = split_target_words(str(entry.get("targetWords", "")))
        row_payload = {
            "row": row,
            "tarih": parse_session_date(str(entry.get("sessionId", ""))),
            "hedefler": words,
            "transcriptWordCount": entry.get("transcriptWordCount"),
            "targetDetectedCount": entry.get("targetDetectedCount"),
            "llmTotalScore": entry.get("llmTotalScore"),
            "transcribeLatencyMs": entry.get("transcribeLatencyMs"),
            "analyzeLatencyMs": entry.get("analyzeLatencyMs"),
            "note": f"auto-import {entry.get('sessionId')}",
        }
        preview_lines.append(
            f"row {row_payload['row']}: {row_payload['tarih']} | "
            f"hedef={', '.join(row_payload['hedefler'])} | "
            f"wc={row_payload['transcriptWordCount']} | "
            f"det={row_payload['targetDetectedCount']} | "
            f"score={row_payload['llmTotalScore']} | "
            f"lat={row_payload['transcribeLatencyMs']}/{row_payload['analyzeLatencyMs']}"
        )

        if not dry_run:
            ws.cell(row, 2).value = row_payload["tarih"]  # Tarih
            for i in range(4):
                ws.cell(row, 4 + i).value = words[i] if i < len(words) else None  # Hedef 1-4
            ws.cell(row, 9).value = row_payload["transcriptWordCount"]  # Toplam Transkript Kelime
            ws.cell(row, 12).value = row_payload["targetDetectedCount"]  # Hedef Tespit Sayisi
            ws.cell(row, 15).value = row_payload["llmTotalScore"]  # LLM Skor
            ws.cell(row, 17).value = row_payload["transcribeLatencyMs"]  # Transcribe ms
            ws.cell(row, 18).value = row_payload["analyzeLatencyMs"]  # Analyze ms
            ws.cell(row, 20).value = row_payload["note"]  # Notlar

        search_start = row + 1
        written += 1

    if preview_lines:
        print("Preview:")
        for line in preview_lines:
            print(f"- {line}")
    return written


def parse_args():
    parser = argparse.ArgumentParser(
        description="Import [voice][qa] logs into VOICE_ECHO_QA.xlsx (QA Oturumları)."
    )
    parser.add_argument("input_path", nargs="?", help="Log file path. If omitted, stdin is used.")
    parser.add_argument(
        "workbook_path",
        nargs="?",
        default="VOICE_ECHO_QA.xlsx",
        help="Target workbook path (default: VOICE_ECHO_QA.xlsx).",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Clear existing manual values in QA rows (2-16) before importing.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview imports without writing workbook.",
    )
    parser.add_argument(
        "--append-from-row",
        type=int,
        default=2,
        help="Start searching for empty rows from this row (2-16, default: 2).",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    input_path = args.input_path
    workbook_path = Path(args.workbook_path)
    append_from_row = args.append_from_row

    if append_from_row < 2 or append_from_row > 16:
        print("--append-from-row must be between 2 and 16.", file=sys.stderr)
        sys.exit(1)

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        sys.exit(1)

    raw = read_text(input_path)
    entries = parse_entries(raw)
    if not entries:
        print("No [voice][qa] entries found.", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(workbook_path)
    if "QA Oturumları" not in wb.sheetnames:
        print("Sheet 'QA Oturumları' not found.", file=sys.stderr)
        sys.exit(1)

    ws = wb["QA Oturumları"]
    if args.overwrite:
        clear_manual_rows(ws)

    written = write_entries(
        ws,
        entries,
        dry_run=args.dry_run,
        append_from_row=append_from_row,
    )
    if args.dry_run:
        print(f"[dry-run] Would write {written} row(s) to {workbook_path}.")
        return

    wb.save(workbook_path)
    print(f"Wrote {written} row(s) to {workbook_path}.")
    if written < len(entries):
        print("Some rows were skipped because the 15-session sheet is full.", file=sys.stderr)


if __name__ == "__main__":
    main()
